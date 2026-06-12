"""
Catálogo dos 72 jogos da fase de grupos, extraído da planilha-modelo.
Lê times e datas (valores em cache) e detecta jogos especiais pela cor verde.

Cada jogo recebe um match_id estável: <grupo><n>, ex.: A1, A2 ... L6.
Esse id é a chave usada em results.csv (resultados oficiais) e nos palpites.

Cache: get_catalog() grava/lê data/catalog.json — assim o robô (GitHub Actions)
funciona sem nenhuma planilha .xlsx no repositório.
"""
import json
import os

import openpyxl
import config as C

_CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "catalog.json")


def _solid_fill(cell):
    f = cell.fill
    if f and f.patternType == "solid":
        rgb = getattr(f.fgColor, "rgb", None)
        return rgb if isinstance(rgb, str) else None
    return None


def build_group_catalog(template_path):
    """Retorna lista de dicts: {match_id, group, home, away, date, sh_cell, sa_cell, special}."""
    wv = openpyxl.load_workbook(template_path, data_only=True)["1a Fase"]   # valores
    ws = openpyxl.load_workbook(template_path)["1a Fase"]                   # estilos (cor)
    matches = []
    for bi, rows in enumerate(C.ROW_BLOCKS):
        for ci, (hcol, scol, sacol, acol, dcol) in enumerate(C.COL_BLOCKS):
            group = C.GROUPS_BY_BLOCK[bi][ci]
            for n, r in enumerate(rows, start=1):
                home = wv[f"{hcol}{r}"].value
                away = wv[f"{acol}{r}"].value
                if not home or not away:
                    continue
                sh_cell, sa_cell = f"{scol}{r}", f"{sacol}{r}"
                special = _solid_fill(ws[sh_cell]) == C.SPECIAL_FILL
                matches.append({
                    "match_id": f"{group}{n}",
                    "group": group,
                    "home": str(home).strip(),
                    "away": str(away).strip(),
                    "date": wv[f"{dcol}{r}"].value,
                    "sh_cell": sh_cell,
                    "sa_cell": sa_cell,
                    "special": special,
                })
    return matches


def get_catalog(template_path=None):
    """Catálogo com cache: usa data/catalog.json se existir; senão lê a planilha
    (se disponível) e grava o cache. É o que o robô usa em CI, sem .xlsx."""
    if template_path is None and os.path.exists(_CACHE):
        with open(_CACHE, encoding="utf-8") as f:
            return json.load(f)
    if template_path is None:
        import glob
        candidates = [f for f in glob.glob(os.path.join(os.path.dirname(_CACHE), "..", "..", "*.xlsx"))
                      if "Modelo v2.1" in f and "~$" not in f]
        if not candidates:
            raise FileNotFoundError(
                "Sem catálogo: nem data/catalog.json nem a planilha-modelo. "
                "Rode 'python3 catalog.py' localmente para gerar o cache.")
        template_path = candidates[0]
    cat = build_group_catalog(template_path)
    os.makedirs(os.path.dirname(_CACHE), exist_ok=True)
    with open(_CACHE, "w", encoding="utf-8") as f:
        json.dump(cat, f, ensure_ascii=False, indent=1)
    return cat


if __name__ == "__main__":
    import sys
    tpl = sys.argv[1] if len(sys.argv) > 1 else None
    cat = get_catalog(tpl)
    print(f"{len(cat)} jogos | especiais: {sum(m['special'] for m in cat)} | cache: {_CACHE}")
    for m in cat:
        star = " ⭐" if m["special"] else ""
        print(f"  {m['match_id']:<4} {m['home']} x {m['away']}  ({m['date']}){star}")
